import openpyxl


def ReadData(file,sheetname,rownum,colnum):
     workbook = openpyxl.load_workbook(file)
     sheet = workbook[sheetname]
     return sheet.cell(row=rownum, column=colnum).value

def WriteData(file, sheetname, rownum, colnumnum, data):
     workbook = openpyxl.load_workbook(file)
     sheet = workbook[sheetname]
     sheet.cell(row=rownum, column=colnumnum).value = data
     workbook.save(file)

# def WriteData(file, sheetname, rownum, colnumnum, data):
#     workbook = openpyxl.load_workbook(file)  # file
#     sheet = workbook[sheetname]  # sheet
#     sheet.cell(row=rownum, column=colnumnum).value = data  # enter data
#     workbook.save(file)  # Save the file


def RowCount(file,sheetname):
     workbook = openpyxl.load_workbook(file)
     sheet = workbook[sheetname]
     return sheet.max_row






